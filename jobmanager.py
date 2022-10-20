#!/usr/bin/env python3
# -*- coding:utf-8 -*-
'''
# File: jobmanager.py
# Project: AUTOJOB
# Created Date: 2022-09-27, 04:16:10
# Author: Chungman Kim(h2noda@unipark.kr)
# Last Modified: Wed Oct 19 2022
# Modified By: Chungman Kim
# Copyright (c) 2022 Unipark
# HISTORY:
# Date      	By	Comments
# ----------	---	----------------------------------------------------------
'''

from asyncio.log import logger

from Sloppy.message import *
from Sloppy.excel import *
from Sloppy.json import *
from Sloppy.common import *
from Sloppy.config import *

import argparse
import os.path
import re

import logging

from openpyxl.utils import get_column_letter
from pkg_resources import split_sections

from rich.progress import track

# RULE_FILE = "./rules.json"
JOB_FIELD = "./jobfield.json"
# CONFIG_FILE = "./config.ini"


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--TYPE", "-T", choices=["J2E", "E2J"],
                        required=True, help="J2E : JIL -> Excel / E2J : Excel -> JIL")
    parser.add_argument("--WRITE", "-W", choices=[
                        "I", "U", "D"], default="I", help="Job Write Type : Insert / Update / Delete")
    parser.add_argument("--JIL", "-J", default="job.jil", help="JIL Filename")
    parser.add_argument("--EXCEL", "-E", default="job.xlsx",
                        help="Excel Filename")

    args = parser.parse_args()

    return args


def initialize(args):

    global now
    global arg_jilfile
    global arg_excelfile
    global arg_write
    global arg_type
    global used_jobfield
    global pattern_jobfield
    global xlsx
    global wb
    global ws

    now = Datetime.get_now("%Y%m%d%H%M%S")

    arg_write = args.WRITE
    arg_type = args.TYPE
    arg_jilfile = args.JIL
    arg_excelfile = args.EXCEL

    try:
        if arg_type == "J2E":
            if os.path.isfile(arg_jilfile) == False:
                raise FileNotFoundError
            else:
                msg.printmsg("Jil File : " + arg_jilfile,
                             "-", 1, "white", False)
                xlsx = Xlsx(arg_excelfile, "JOB", mode="n")
                wb = xlsx.wb
                ws = xlsx.ws
        elif arg_type == "E2J":
            if os.path.isfile(arg_excelfile) == False:
                raise FileNotFoundError
            else:
                msg.printmsg("Excel File : " + arg_excelfile,
                             "-", 1, "white", False)
                xlsx = Xlsx(arg_excelfile, "JOB", mode="r")
                wb = xlsx.wb
                ws = xlsx.ws

        used_jobfield = read_used_jobfield(JOB_FIELD)
        pattern_jobfield = read_pattern_jobfield(JOB_FIELD)
        # cfg_data = Config()
    except FileNotFoundError:
        msg.printmsg("Error : Convert File Not Found!!", "", 0, "red", True)
        exit()


def read_used_jobfield(jobfield):
    job_field = Json.read_jsonfile(jobfield)
    used_field = {}
    used_field = [i["field"] for i in job_field["field"] if i["used"] == "Y"]
    return used_field


def read_pattern_jobfield(jobfield):
    job_field = Json.read_jsonfile(jobfield)
    pattern_jobfield = {}
    pattern_jobfield = [i["field"]
                        for i in job_field["field"] if (i["used"] == "Y" and i["pattern"] == "Y")]
    return pattern_jobfield


def count_jillines(filename):
    with open(filename, "r") as infile:
        i = -1
        for i, _ in enumerate(infile):
            pass
        return i + 1


def convert_j2e():

    global count

    ws.append(used_jobfield)

    linecount = count_jillines(arg_jilfile)
    count = 0
    list_of_alljob = []
    job = {}    # 개별 Job 정보
    with open(arg_jilfile, "r") as jil:
        for line_jil in track(jil, description="[white]  - COnvert Job Info :", total=linecount):
            if "insert_job:" in line_jil:
                list_of_alljob.append(job)
                line_jil = line_jil.strip()
                jobName = re.findall(
                    r"insert_job:(.*?)job_type:", line_jil)[0]
                jobType = line_jil.split("job_type:")[1]
                job = {}
                job["insert_job"] = str(jobName).strip()
                job["job_type"] = str(jobType).strip()
                count += 1
            else:
                line_jil = line_jil.strip()
                if line_jil != "\n" and "/* ----" not in line_jil and line_jil != "":
                    if "start_times" in line_jil:
                        spli = line_jil.split("start_times:")
                        job["start_times"] = str(
                            spli[1]).replace("\"", "").strip()
                    elif "permission:" in line_jil:
                        spli = line_jil.split("permission:")
                        spli_value = str(spli[1]).strip()
                        if spli_value:
                            job["permission"] = str(spli[1]).strip()
                        else:
                            job["permission"] = ""
                    elif "command:" in line_jil:
                        spli = line_jil.split("command:")
                        job["command"] = str(spli[1]).strip()
                    # notification_emailaddress 처리
                    elif "notification_emailaddress" in line_jil:
                        spli = line_jil.split(
                            "notification_emailaddress:", 1)
                        if "notification_emailaddress" not in job:
                            job["notification_emailaddress"] = str(
                                spli[1]).strip()
                        else:
                            job["notification_emailaddress"] += ", " + \
                                str(spli[1]).strip()
                    else:
                        spli = line_jil.split(":", 1)
                        job[str(spli[0]).strip()] = str(
                            spli[1]).strip().replace("\"", "")

        list_of_alljob.append(job)

    list_of_alljob.pop(0)
    for ar in list_of_alljob:
        values = []
        for k in used_jobfield:
            if k in ar:
                values.append(ar[k])
            else:
                values.append(None)
        ws.append(values)
        count += 1

    msg.printmsg("Save Excel File", "-", 1, "white", False)

    # Resize - Excel File Column Width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        # print(column)
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 1) * 1.1
        if adjusted_width > 50:
            ws.column_dimensions[get_column_letter(column)].width = 50
        else:
            ws.column_dimensions[get_column_letter(
                column)].width = adjusted_width
    wb.save(arg_excelfile)


def convert_e2j():
    max_row = xlsx.ws.max_row
    max_col = xlsx.ws.max_column

    row_header = xlsx.ws[1]
    row_range = xlsx.ws[2:max_row]

    with open(arg_jilfile, "w") as jilfile:
        for rows in track(row_range, description="[white]  - Convert Job Info :", total=max_row):
            idx_column = 0
            jilfile.write("\n\n/* " + "-" * 17 + " " +
                          str(rows[0].value + " " + "-" * 17 + " */\n\n"))

            for cell in rows:
                header = str(row_header[idx_column].value)

                if cell.value is not None:
                    if header in pattern_jobfield:
                        celldata = "\"" + cell.value + "\""
                    else:
                        celldata = cell.value
                    if (header == "insert_job" or header == "updatE_job"):
                        jilfile.write(header + ": " + celldata + "   ")
                    else:
                        if header == "permission":  # "permission" 필드 권한 처리
                            if celldata == "gx,ge,wx,we,mx,me":
                                jilfile.write(header + ": \n")
                            else:
                                jilfile.write(header + ": " + celldata + "\n")
                        elif header == "notification_emailaddress":
                            splited_celldata = celldata.split(",")
                            for sp_cell in splited_celldata:
                                jilfile.write(header + ": " +
                                              sp_cell.strip() + "\n")
                        else:
                            jilfile.write(header + ": " + celldata + "\n")

                idx_column += 1
    msg.printmsg("Convert Job : " + str(max_row), "-", 1)
    msg.printmsg("Finished....", "-", 1)


def main():
    global logger

    logger = logging.getLogger()

    try:
        args = get_args()
        # Step 00. Title
        msg.title("Autosys Job Converter", color="blue")

        # Step 01. Initialize(Load Rule / Config, Env, etc)
        msg.printmsg("Step #01. Initialize...", "", 0, "green", True)

        msg.printmsg("Load Environment variable", "-", 1)
        initialize(args)

        # Step 02. Read File
        # msg.printmsg("Step #02. JIL / Excel File Read", "", 0, "green", True)

        # Step 03. Convert
        msg.printmsg("Step #02. Convert...", "", 0, "green", True)
        if args.TYPE == "J2E":
            # Step 02-01. Convert Jil to Excel
            msg.printmsg("Convert Type : JIL to Excel", "-", 1, "white", False)
            convert_j2e()
        elif args.TYPE == "E2J":
            # Step 02-02. Convert Excel to Jil
            msg.printmsg("Convert Type : Excel to JIL", "-", 1, "white", False)
            convert_e2j()

    # pass

    # except:
    #     msg.blank()
    # msg.printmsg("Error...Error...Error...", "", 0, "red", True)
    # pass

    except Exception as e:
        logger.exception(str(e))
        msg.printmsg("Error...Error...Error...", "", 0, "red", True)
    else:
        # Step 99. Save
        # pass
        msg.OK()

    finally:
        msg.blank()
        msg.copyright()
        msg.blank()


if __name__ == "__main__":

    msg = Msg()
    main()
