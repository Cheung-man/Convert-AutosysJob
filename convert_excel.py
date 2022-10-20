#!/usr/bin/env python3
# -*- coding:utf-8 -*-
'''
# File: jobmanager.py
# Project: AUTOJOB
# Created Date: 2022-09-27, 04:16:10
# Author: Chungman Kim(h2noda@unipark.kr)
# Last Modified: Thu Oct 20 2022
# Modified By: Chungman Kim
# Copyright (c) 2022 Unipark
# HISTORY:
# Date      	By	Comments
# ----------	---	----------------------------------------------------------
'''

from asyncio.log import logger

from Sloppy.message import *
from Sloppy.excel import *
from Sloppy.json import *
from Sloppy.common import *

import argparse
import os.path
import re

import logging

from openpyxl.utils import get_column_letter
from rich.progress import track

JOB_FIELD = "./jobfield.json"


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--EXCEL", "-E", help="Excel Filename")
    args = parser.parse_args()

    return args


def convert_filename(filename, filetype):
    """파일의 확장자를 변경

    Args:
        filename (String): 대상 파일명
        filetype (String): 바뀔 확장자명

    Returns:
        String: 변경된 파일명
    """
    name, ext = os.path.splitext(filename)
    retval = name + "." + filetype
    return retval


def initialize(args):

    global arg_jilfile
    global arg_excelfile
    global used_jobfield
    global dq_jobfield
    global xlsx
    global wb
    global ws

    arg_excelfile = args.EXCEL
    arg_jilfile = convert_filename(arg_excelfile, "jil")

    try:
        if os.path.isfile(arg_excelfile) == False:
            raise FileNotFoundError
        else:
            msg.printmsg("Excel File : " + arg_excelfile,
                         "->", 1, "white", False)
            xlsx = Xlsx(arg_excelfile, "JOB", mode="r")
            wb = xlsx.wb
            ws = xlsx.ws

        used_jobfield = read_used_jobfield(JOB_FIELD)
        dq_jobfield = read_dq_jobfield(JOB_FIELD)
    except FileNotFoundError:
        msg.printmsg("Error : Convert File Not Found!!", "", 0, "red", True)
        exit()


def read_used_jobfield(jobfield):
    """Job 정보에서 사용될 Field를 읽음
        - "used" 값이 "Y" : Excel 파일에 기록

    Args:
        jobfield (String): 사용될 Job Field

    Returns:
        _type_: _description_
    """
    job_field = Json.read_jsonfile(jobfield)
    used_field = {}
    used_field = [i["field"] for i in job_field["field"] if i["used"] == "Y"]
    return used_field


def read_dq_jobfield(jobfield):
    """Double Quotation 사용 필드 설정

    Args:
        jobfield (String): Job Field

    Returns:
        _type_: _description_
    """
    job_field = Json.read_jsonfile(jobfield)
    dq_jobfield = {}
    dq_jobfield = [i["field"]
                   for i in job_field["field"] if (i["used"] == "Y" and i["double_quotation"] == "Y")]
    return dq_jobfield


def convert_e2j():
    max_row = xlsx.ws.max_row
    cnt_job = max_row - 1

    row_header = xlsx.ws[1]
    row_range = xlsx.ws[2:max_row]

    with open(arg_jilfile, "w") as jilfile:
        for rows in track(row_range, description="[white]  - Convert Job Info :", total=cnt_job):
            idx_column = 0
            jilfile.write("\n\n/* " + "-" * 17 + " " +
                          str(rows[0].value + " " + "-" * 17 + " */\n\n"))

            for cell in rows:
                header = str(row_header[idx_column].value)

                if cell.value is not None:
                    if header in dq_jobfield:
                        celldata = "\"" + cell.value + "\""
                    else:
                        celldata = cell.value
                    if (header == "insert_job" or header == "updatE_job"):
                        jilfile.write(header + ": " + celldata + "   ")
                    else:
                        if header == "notification_emailaddress":
                            splited_celldata = celldata.split(",")
                            for sp_cell in splited_celldata:
                                jilfile.write(header + ": " +
                                              sp_cell.strip() + "\n")
                        else:
                            jilfile.write(header + ": " + celldata + "\n")
                else:
                    if header == "permission":
                        jilfile.write(header + ": \n")

                idx_column += 1

    msg.printmsg("Convert Job : " + str(cnt_job), "->", 1)


def main():
    global logger

    logger = logging.getLogger()

    try:
        args = get_args()
        # Step 00. Title
        msg.title("   Autosys Job Converter   ", color="blue")

        # Step 01. Initialize(Load Rule / Config, Env, etc)
        msg.printmsg("Step #01. Initialize...", "", 0, "green", True)

        msg.printmsg("Load Environment variable", "->", 1)
        initialize(args)

        # Step 02. Convert
        msg.printmsg("Step #02. Convert...", "", 0, "green", True)
        msg.printmsg("Convert Type : Excel to JIL", "->", 1, "white", False)
        convert_e2j()
        msg.printmsg("Finished....", "->", 1)

    except Exception as e:
        logger.exception(str(e))
        msg.printmsg("Error...Error...Error...", "", 0, "red", True)
    else:
        # Step 99. OK
        msg.OK()

    finally:
        msg.blank()
        msg.copyright()
        msg.blank()


if __name__ == "__main__":

    msg = Msg()
    main()
